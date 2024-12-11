import openpyxl

def find_device_port(room_number, worksheet_name):
    # Open the workbook and select the worksheet
    workbook = openpyxl.load_workbook('portAllocation.xlsx')
    worksheet = workbook[worksheet_name]

    TVport = None
    
    # Iterate through rows to find the cell with room number followed by "TV"
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value == room_number:
                room_cell = cell
                # Check if the next cell in the same row contains "TV"
                if room_cell.col_idx < len(row) and row[room_cell.col_idx].value == "TV":
                    TVport = worksheet.cell(row=room_cell.row, column=room_cell.col_idx-1).value
                    break
        if TVport:
            break

    if not TVport:
        return "Device or TV port not found"

    # Extract the number from the D_value
    if TVport and TVport.startswith('D'):
        number = int(TVport[1:])
    else:
        return "Invalid D_value"

    # Determine the switch and port
    result = {}
    if number <= 48:
        result["switch"] = 1
        result["port"] = number
    elif 49 <= number <= 96:
        result["switch"] = 2
        result["port"] = number - 48
    else:
        result["switch"] = 3
        result["port"] = number - 96

    return result

# Example usage
room_number = 722
worksheet_name = "7th floor "
port_info = find_device_port(room_number, worksheet_name)
print(port_info)
